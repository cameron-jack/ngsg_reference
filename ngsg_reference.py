#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
@created: 12 Dec 2024
@author: Cameron Jack, ANU Bioinformatics Consultancy,
        JCSMR, Australian National University

Streamlit application to read an Excel XLSX two column spreadsheet
containing names and DNA sequences respectively, and output as FASTA.

        Run with: run_reference.bat
"""
#from csv import field_size_limit
#from msilib.schema import File
#from xmlrpc.client import boolean
#import jsonpickle
#import os
#import sys
from pathlib import Path  
#from subprocess import check_output, CalledProcessError, STDOUT
#import subprocess
import pandas as pd
import chardet  # For automatic encoding detection
import streamlit as st
from openpyxl import load_workbook

import bin.util as util
from stutil import hline, add_css
#import extra_streamlit_components as stx
#from time import sleep

###
# Use Eslam Ibrahim's code below to check for non-ASCII characters
###

######################################################################################################
# This script is used to validate the format of FASTA files.
# It is designed to be used in NGS genotyping pipelines.
# The script utilizes Streamlit to provide a web interface for the end-user.
# To run the script, use the command: #"python -m streamlit run fasta_checker.py"
# This script was developed and tested by Eslam Ibrahim (Eslam.ibrahim@anu.edu.au) on OCT. 10, 2024.
######################################################################################################

def check_non_ascii(file_content, issues):
    """
    A function to check for non-ASCII characters.
    Inputs: character stream, issues passed by reference
    Outputs: list of dictionaries
    """
    lines = file_content.splitlines()

    for line_num, line in enumerate(lines):
        for idx, char in enumerate(line):
            if ord(char) > 127:  # ASCII characters have values from 0 to 127
                issues.append({
                    'Issue Number': len(issues)+1,
                    'Line Number': line_num,
                    'Issue': f"Invalid character in sequence header: '{char}' at position {idx + 1}"
                })
        line_num += 1  # Move to the next line number


def check_valid_sequence(file_content, issues):
    """
    Function to check that sequences contain only A, T, C, G, N
    Extended to allow parentheses for variable sections
    Spaces and tabs will not be reported here, only checked in check_gaps
    Inputs: character stream, issues list passed by reference
    """
    lines = file_content.splitlines()

    for line_num,line in enumerate(lines):
        if line.startswith(">") or not line.strip():  # Ignore headers and blank lines
            continue
        else:
            sequence = line.strip().replace(" ", "").replace("\t", "")  # Clean the sequence
            #print(sequence)
            for idx, char in enumerate(sequence):
                if char not in "AaTtCcGgNn()":  # Check for valid characters
                    issues.append({
                        'Issue Number': len(issues)+1,
                        'Line Number': line_num+1,
                        'Issue': f"Invalid character in sequence: '{char}' at position {idx + 1}"
                    })


def check_gaps(file_content, issues):
    """
    Function to check for gaps (spaces or tabs) within sequences.
    Inputs: character stream, issues list passed by reference
    """
    lines = file_content.splitlines()

    for line_num,line in enumerate(lines):
        if line.startswith(">") or not line.strip():  # Ignore headers and blank lines
            continue
        else:
            if " " in line or "\t" in line:
                issues.append({
                    'Issue Number': len(issues)+1,
                    'Line Number': line_num+1,
                    'Issue': "Gap (space or tab) in sequence"
                })


def check_blank_lines(file_content, issues):
    """
    FASTA files should not contain blank lines
    Inputs: character stream, issues list passed by reference
    Outputs: None
    """
    lines = file_content.splitlines()

    for line_num, line in enumerate(lines):
        if not line.strip():  # Blank line
            issues.append({
                'Issue Number': len(issues)+1,
                'Line Number': line_num+1,
                'Issue': "Blank line found"
            })


def check_fasta_file(file_content):
    """
    Issue feature checks for FASTA specific file features
    Inputs: character stream
    Outputs: list of dictionaries
    """
    issues = []
    # pass issues by reference
    check_non_ascii(file_content, issues)
    check_valid_sequence(file_content, issues)
    check_gaps(file_content, issues)
    check_blank_lines(file_content, issues)
    return issues   


def display_file_character_issues(issues):
    """
    Common GUI widget for displaying any file check problems
    """
    # Create DataFrame and sort by Line Number if necessary
    df = pd.DataFrame(issues)
    # df = df.sort_values(by="Line Number")  # Uncomment if you want to sort by line number
        
    # Display the number of issues found and the table with custom size
    st.write(f"There are {len(issues)} issues found in the file:")
    st.dataframe(df, height=300, width=500, hide_index=True)


#=====================================
# Streamlit UI

def main():
    """
    The NGSgeno "Reference" application. Converts two-column (name, sequence) XLSX to FASTA
    """    
    st.set_page_config(
        page_title="NGSG Reference converter",
        page_icon="ngsg_icon.png",
        layout="wide"
    )
    add_css()
    st.title("NGS Genotyping: Reference converter")
    st.subheader("Upload an Excel document containing references in two columns (name and sequence)")
    hline()
    okay_to_save = False
    format_as_fasta = False
    table = []
    legit_file = True
    with st.container(border=True):
        uploaded_file = st.file_uploader("Choose an Excel reference sequence file (XLSX format)")
        if uploaded_file is not None:
            fname = uploaded_file.name
            #tmp_fp = Path('tmp.xlsx')
            #if tmp_fp.exists():
            #    tmp_fp.unlink()
            #with open(tmp_fp, 'wb') as fout:
            #    fout.write(uploaded_file)
            try:
                wb = load_workbook(uploaded_file, read_only=True, keep_links=False)
            except Exception as exc:
                st.error(f'File does not appear to be in Excel XLSX format. Please try again with another file.')
                #st.error(f'{exc}')  # the exception messages aren't very helpful
                legit_file = False
            if legit_file:
                sheet = wb.active
                table = [list(map(str,cells)) for cells in sheet.iter_rows(values_only=True, min_col=1,max_col=2)]
                illegal_chars_seen = False
                hline()
                for i, row in enumerate(table):
                    for j, col in enumerate(row):
                        for k, c in enumerate(col):
                            if ord(c) > 255:
                                st.write(f'Non-ASCII character in row:{i+1} column:{j+1} character:{k+1}, symbol:{ord(c)} {c} in {col}')
                                illegal_chars_seen = True
                if illegal_chars_seen:
                    st.write('')
                    st.write('Please correct illegal characters and upload again')
                else:
                    format_as_fasta = True

    fasta_contents = []
    if format_as_fasta and table:
        header_blank = False
        for row in table:
            for j, col in enumerate(row):
                txt = col.strip()
                if txt.lower() == 'none':
                    txt = ''
                #if not txt or txt.lower() == 'none':
                #    continue
                if j % 2 == 0:
                    if not txt:
                        header_blank = True
                    else:
                        header_blank = False
                        name_line = '>' + txt
                        fasta_contents.append(name_line)
                        #fout.write('>' + col.strip())
                else:
                    if not header_blank:
                        fasta_contents.append(txt)
                        #fout.write(col.strip())
        fasta_str = '\n'.join(fasta_contents)
        # check the formating before we allow the FASTA to be written to file
        if fasta_contents:
            issues = check_fasta_file(fasta_str)
            if not issues:
                okay_to_save = True
            else:
                display_file_character_issues(issues)

    if okay_to_save:
        do_save = False
        save_path = Path(f'../Downloads/{fname.replace(".xlsx",".fa")}')
        if Path(save_path).exists():
            st.warning(f'File already exists with name {save_path}')
            save_button = st.button('Overwrite FASTA')
            if save_button:
                do_save = True
                st.write('Saving to FASTA format')
        else:
            do_save = True
        if do_save:
            with open(Path(save_path), 'wt') as fout:
                fout.write(fasta_str)
                #for row in table:
                #    for j, col in enumerate(row):
                #        if j % 2 == 0:
                #            fout.write('>' + col.strip())
                #        else:
                #            fout.write(col.strip())
            st.write(f'FASTA file saved to {save_path}')
            table = []
            fasta_contents = []


if __name__ == '__main__':
    main()